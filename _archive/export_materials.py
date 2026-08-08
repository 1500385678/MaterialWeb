"""
导出 MaterialDB 全部材料为 Excel 模板格式
用法: python export_materials.py [输出文件名.xlsx]
"""
import os
import sys
import json
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

DB = r'D:\Mac\Mac\workteam\05_space\03_architect\_ArchitectLib\MaterialDb\materials.db'
OUT = sys.argv[1] if len(sys.argv) > 1 else '材料库_导出.xlsx'

# 与 import_materials.py / generate_template.py 完全一致的列定义
COLUMNS = [
    ('编码',          'code',             'str'),
    ('中文名',        'name_cn',          'str'),
    ('英文名',        'name_en',          'str'),
    ('分类代码',      'category_code',    'str'),
    ('子分类',        'sub_category',     'str'),
    ('单价',          'unit_price',       'num'),
    ('单位',          'unit',             'str'),
    ('单价(立方)',    'unit_price_m3',    'num'),
    ('施工费',        'labor_cost',       'num'),
    ('损耗系数',      'loss_factor',      'num'),
    ('防火等级',      'fire_rating',      'str'),
    ('防火说明',      'fire_note',        'str'),
    ('造价档',        'cost_tier',        'str'),
    ('环保等级',      'env_grade',        'str'),
    ('规范标准',      'std_code',         'str'),
    ('环保认证',      'eco_cert',         'str'),
    ('密度',          'density',          'str'),
    ('强度',          'strength',         'str'),
    ('导热系数',      'thermal_cond',     'str'),
    ('吸水率',        'water_absorp',     'str'),
    ('质感',          'texture',          'list'),
    ('色系',          'color_series',     'list'),
    ('常见规格',      'specs',            'str'),
    ('肌理',          'patterns',         'str'),
    ('应用场景',      'applications_json','list'),
    ('视觉效果',      'visual_desc',      'str'),
    ('构造节点',      'structure_notes',  'str'),
    ('耐久性',        'durability',       'str'),
    ('使用寿命',      'lifespan_years',   'str'),
    ('维护周期',      'maintenance',      'str'),
    ('供应商',        'suppliers_json',   'str'),
    ('图片文件名',    'image_urls',       'list'),
    ('考试要点',      'exam_points',      'list'),
    ('考试案例',      'exam_cases',       'str'),
    ('考试权重',      'exam_weight',      'num'),
    ('来源',          'source_doc',       'str'),
    ('备注',          'remark',           'str'),
]

# JSON 字段转换方向：DB 存 JSON 字符串，Excel 用 | 分隔
JSON_TO_EXCEL = {
    'applications_json': 'applications',
    'suppliers_json':     'suppliers',
}

HEADER_FILL = PatternFill('solid', fgColor='FF9E4A')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
THIN = Side(style='thin', color='999999')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center')

# 与 generate_template.py 一致的列宽
WIDTHS = {
    '编码': 12, '中文名': 14, '英文名': 22, '分类代码': 14, '子分类': 16,
    '单价': 10, '单位': 8, '单价(立方)': 12, '施工费': 10, '损耗系数': 10,
    '防火等级': 10, '防火说明': 16, '造价档': 8, '环保等级': 14, '规范标准': 16,
    '环保认证': 14, '密度': 18, '强度': 20, '导热系数': 16, '吸水率': 10,
    '质感': 20, '色系': 16, '常见规格': 22, '肌理': 18, '应用场景': 24,
    '视觉效果': 36, '构造节点': 30, '耐久性': 10, '使用寿命': 12, '维护周期': 18,
    '供应商': 20, '图片文件名': 28, '考试要点': 28, '考试案例': 22, '考试权重': 10,
    '来源': 22, '备注': 28,
}
# 列名 → Excel 显示名（与 COLUMNS 对应）
COL_TIPS = {
    '编码': '唯一标识，如 STONE_005',
    '中文名': '材质中文名',
    '英文名': '材质英文名',
    '分类代码': '见"字典"sheet 的分类代码',
    '子分类': '细分类型，如"天然石材"',
    '单价': '元/m²',
    '单位': '默认 m²，可改 m/张/块',
    '单价(立方)': '元/m³',
    '施工费': '元/m²',
    '损耗系数': '默认 1.05',
    '防火等级': 'A1 / B1 / B2',
    '防火说明': '防火补充说明',
    '造价档': '低 / 中 / 高',
    '环保等级': '环保认证等级',
    '规范标准': '如 GB/T 18601',
    '环保认证': '如 中国环境标志',
    '密度': '如 2500~2800 kg/m³',
    '强度': '如 抗压 100~250MPa',
    '导热系数': '如 2.0~3.5 W/(m·K)',
    '吸水率': '如 <0.5%',
    '质感': '用 | 分隔，如 光滑|哑光|粗糙',
    '色系': '用 | 分隔，如 灰|米黄|红',
    '常见规格': '如 600×600 / 800×800',
    '肌理': '如 光面/火烧面/荔枝面',
    '应用场景': '用 | 分隔，如 外墙|幕墙|室内',
    '视觉效果': '一两句话描述视觉效果',
    '构造节点': '构造做法说明',
    '耐久性': '高/中/低',
    '使用寿命': '如 50~100 年',
    '维护周期': '如 5~10年做一次防护处理',
    '供应商': '用 | 分隔名字，如 环球石材|康利石材',
    '图片文件名': '用 | 分隔，放到 media/images/',
    '考试要点': '用 | 分隔',
    '考试案例': '一句话',
    '考试权重': '0~1 之间',
    '来源': '信息出处',
    '备注': '自由文本',
}


def json_to_pipe(v):
    """把 DB 里存的 JSON 数组字符串转成 'a|b|c'，空/无值返回空串"""
    if not v: return ''
    try:
        arr = json.loads(v)
        if isinstance(arr, list):
            return '|'.join(str(x) for x in arr if str(x).strip())
    except Exception:
        return str(v)
    return ''


def main():
    if not os.path.exists(DB):
        print(f'✗ 数据库不存在: {DB}')
        sys.exit(1)

    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.row_factory = sqlite3.Row
    rows = c.execute("""
        SELECT m.*, c.code AS cat_code
        FROM materials m
        LEFT JOIN categories c ON c.id = m.category_id
        WHERE m.status = 'active'
        ORDER BY m.code
    """).fetchall()
    conn.close()
    print(f'>>> DB 里有 {len(rows)} 张材质')

    wb = Workbook()
    ws = wb.active
    ws.title = '材料'

    # 表头
    col_names = [n for n, _, _ in COLUMNS]
    for col_idx, name in enumerate(col_names, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        if name in COL_TIPS:
            cell.comment = Comment(COL_TIPS[name], 'MaterialWeb')

    # 数据行
    for r_idx, row in enumerate(rows, 2):
        rec = dict(row)
        for col_idx, (name, db_field, ftype) in enumerate(COLUMNS, 1):
            if db_field == 'category_code':
                val = rec.get('cat_code') or ''
            elif ftype == 'list':
                val = json_to_pipe(rec.get(db_field))
            else:
                val = rec.get(db_field)
                if val is None: val = ''
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.alignment = WRAP
            cell.border = BORDER

    # 列宽
    for col_idx, (name, _, _) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = WIDTHS.get(name, 14)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'D2'

    # ============================
    # Sheet 字典
    # ============================
    ws2 = wb.create_sheet('字典')
    for col_idx, h in enumerate(['分类代码', '分类名称', '父级'], 1):
        c = ws2.cell(row=1, column=col_idx, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER
    CATS = [
        ('metal', '金属材料', ''), ('metal.steel', '建筑钢材', 'metal'),
        ('metal.aluminum', '铝合金', 'metal'), ('metal.copper', '铜/钛/锌', 'metal'),
        ('metal.stainless', '不锈钢', 'metal'),
        ('concrete', '混凝土', ''), ('concrete.normal', '普通混凝土', 'concrete'),
        ('concrete.hpc', '高性能混凝土', 'concrete'),
        ('concrete.scc', '自密实混凝土', 'concrete'),
        ('concrete.fairface', '清水混凝土', 'concrete'),
        ('masonry', '砌体材料', ''), ('wood', '木材', ''),
        ('glass', '玻璃', ''), ('glass.float', '浮法玻璃', 'glass'),
        ('glass.tempered', '钢化玻璃', 'glass'), ('glass.laminated', '夹层玻璃', 'glass'),
        ('glass.low_e', 'Low-E中空', 'glass'), ('glass.vacuum', '真空玻璃', 'glass'),
        ('stone', '石材', ''), ('stone.granite', '花岗岩', 'stone'),
        ('stone.marble', '大理石', 'stone'), ('stone.sandstone', '砂岩', 'stone'),
        ('stone.artificial', '人造石材', 'stone'),
        ('membrane', '膜材', ''), ('insulation', '保温/防水', ''),
        ('finishing', '装饰材料', ''), ('finishing.paint', '涂料', 'finishing'),
        ('finishing.tile', '瓷砖', 'finishing'), ('finishing.metal', '金属板', 'finishing'),
        ('finishing.wood', '木饰面', 'finishing'), ('finishing.glass', '玻璃', 'finishing'),
        ('composite', '复合材料', ''),
    ]
    for i, (code, name, parent) in enumerate(CATS, 2):
        ws2.cell(row=i, column=1, value=code).border = BORDER
        ws2.cell(row=i, column=2, value=name).border = BORDER
        ws2.cell(row=i, column=3, value=parent).border = BORDER
    row = len(CATS) + 4
    for col_idx, h in enumerate(['防火等级', '含义'], 1):
        c = ws2.cell(row=row, column=col_idx, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER
    for i, (k, v) in enumerate([('A1','不燃材料'),('B1','难燃材料'),('B2','可燃材料')], 1):
        ws2.cell(row=row+i, column=1, value=k).border = BORDER
        ws2.cell(row=row+i, column=2, value=v).border = BORDER
    row += 5
    for col_idx, h in enumerate(['造价档', '范围'], 1):
        c = ws2.cell(row=row, column=col_idx, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER
    for i, (k, v) in enumerate([('低','<200 元/m²'),('中','200~500 元/m²'),('高','>500 元/m²')], 1):
        ws2.cell(row=row+i, column=1, value=k).border = BORDER
        ws2.cell(row=row+i, column=2, value=v).border = BORDER
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 14

    # ============================
    # Sheet 说明
    # ============================
    ws3 = wb.create_sheet('说明')
    ws3.cell(row=1, column=1, value='材料表录入说明').font = Font(bold=True, size=14, color='FF9E4A')
    NOTES = [
        '', '【必填项】编码、中文名、分类代码。其他都可留空。',
        '', '【分类代码】从"字典"sheet 选（如 masonry / metal）。',
        '', '【多值字段】用 | 分隔：',
        '   质感: 哑光|粗糙|自然纹理', '   应用场景: 外墙|幕墙|室内',
        '   图片文件名: aaa.jpg|bbb.jpg', '', '【图片】放到 media/images/ 目录下，文件名写进"图片文件名"列。',
        '', '【编码】要唯一。导入时编码已存在 → 默认跳过，加 --overwrite 覆盖。',
        '', '【使用流程】',
        '   1. 在"材料"sheet 编辑（不要改"字典"和"说明"）',
        '   2. 保存文件', '   3. python import_materials.py 本文件名.xlsx',
        '   4. 刷 MaterialWeb 页面', '', '【干跑】python import_materials.py 文件.xlsx --dry-run',
        '', '【冲突策略】', '   默认跳过（提示）',
        '   --overwrite 覆盖', '   --strict 报错退出',
    ]
    for i, line in enumerate(NOTES, 2):
        ws3.cell(row=i, column=1, value=line).alignment = WRAP
    ws3.column_dimensions['A'].width = 80

    wb.save(OUT)
    print(f'✓ 已导出: {os.path.abspath(OUT)}')
    print(f'  共 {len(rows)} 行材质数据')


if __name__ == '__main__':
    main()
