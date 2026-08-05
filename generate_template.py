"""
生成 MaterialDB 材料录入 Excel 模板
用法: python generate_template.py [输出文件名.xlsx]
"""
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

OUT = sys.argv[1] if len(sys.argv) > 1 else '材料表_模板.xlsx'

wb = Workbook()

# 通用样式
HEADER_FILL = PatternFill('solid', fgColor='FF9E4A')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
EXMAPLE_FILL = PatternFill('solid', fgColor='FFF4E6')
THIN = Side(style='thin', color='999999')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center')

# ============================
# Sheet 1: 材料（主表）
# ============================
ws = wb.active
ws.title = '材料'

COLUMNS = [
    # (列名, DB 字段, 必填, 类型, 说明)
    ('编码',          'code',             True,  'str',  '唯一标识，如 STONE_005'),
    ('中文名',        'name_cn',           True,  'str',  '材质中文名'),
    ('英文名',        'name_en',           False, 'str',  '材质英文名'),
    ('分类代码',      'category_code',     True,  'str',  '见"字典"sheet 的分类代码'),
    ('子分类',        'sub_category',      False, 'str',  '细分类型，如"天然石材"'),
    ('单价',          'unit_price',        False, 'num',  '元/m²'),
    ('单位',          'unit',              False, 'str',  '默认 m²，可改 m/张/块'),
    ('单价(立方)',    'unit_price_m3',     False, 'num',  '元/m³'),
    ('施工费',        'labor_cost',        False, 'num',  '元/m²'),
    ('损耗系数',      'loss_factor',       False, 'num',  '默认 1.05'),
    ('防火等级',      'fire_rating',       False, 'str',  'A1 / B1 / B2'),
    ('防火说明',      'fire_note',         False, 'str',  '防火补充说明'),
    ('造价档',        'cost_tier',         False, 'str',  '低 / 中 / 高'),
    ('环保等级',      'env_grade',         False, 'str',  '环保认证等级'),
    ('规范标准',      'std_code',          False, 'str',  '如 GB/T 18601'),
    ('环保认证',      'eco_cert',          False, 'str',  '如 中国环境标志'),
    ('密度',          'density',           False, 'str',  '如 2500~2800 kg/m³'),
    ('强度',          'strength',          False, 'str',  '如 抗压 100~250MPa'),
    ('导热系数',      'thermal_cond',      False, 'str',  '如 2.0~3.5 W/(m·K)'),
    ('吸水率',        'water_absorp',      False, 'str',  '如 <0.5%'),
    ('质感',          'texture',           False, 'list', '用 | 分隔，如 光滑|哑光|粗糙'),
    ('色系',          'color_series',      False, 'list', '用 | 分隔，如 灰|米黄|红'),
    ('常见规格',      'specs',             False, 'str',  '如 600×600 / 800×800'),
    ('肌理',          'patterns',          False, 'str',  '如 光面/火烧面/荔枝面'),
    ('应用场景',      'applications',      False, 'list', '用 | 分隔，如 外墙|幕墙|室内'),
    ('视觉效果',      'visual_desc',       False, 'str',  '一两句话描述视觉效果'),
    ('构造节点',      'structure_notes',   False, 'str',  '构造做法说明'),
    ('耐久性',        'durability',        False, 'str',  '高/中/低'),
    ('使用寿命',      'lifespan_years',    False, 'str',  '如 50~100 年'),
    ('维护周期',      'maintenance',       False, 'str',  '如 5~10年做一次防护处理'),
    ('供应商',        'suppliers',         False, 'str',  '用 | 分隔名字，如 环球石材|康利石材'),
    ('图片文件名',    'image_urls',        False, 'list', '用 | 分隔，放到 media/images/'),
    ('考试要点',      'exam_points',       False, 'list', '用 | 分隔'),
    ('考试案例',      'exam_cases',        False, 'str',  '一句话'),
    ('考试权重',      'exam_weight',       False, 'num',  '0~1 之间'),
    ('来源',          'source_doc',        False, 'str',  '信息出处'),
    ('备注',          'remark',            False, 'str',  '自由文本'),
]

# 写表头
for col_idx, (name, _, _, _, tip) in enumerate(COLUMNS, 1):
    cell = ws.cell(row=1, column=col_idx, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER
    # 加 comment 提示
    cell.comment = Comment(tip, 'MaterialWeb')

# 一行示例（让用户能直接跑通流程）
EXAMPLE = [
    'FLEX_001',     # 编码
    '软瓷片',        # 中文名
    'Flexible Ceramic Tile',  # 英文名
    'masonry',      # 分类代码
    '柔性饰面砖',     # 子分类
    180,            # 单价
    'm²',           # 单位
    '',             # 单价(立方)
    80,             # 施工费
    1.05,           # 损耗
    'A1',           # 防火
    '不燃材料',       # 防火说明
    '中',           # 造价
    '中国环境标志',     # 环保
    'JG/T 252',     # 规范
    '',             # 环保认证
    '1700 kg/m³',   # 密度
    '抗折≥15MPa',     # 强度
    '0.5 W/(m·K)',  # 导热
    '<3%',          # 吸水
    '哑光|粗糙|自然纹理',  # 质感
    '灰|米黄|红棕',     # 色系
    '240×60 / 300×150，厚3~5mm',  # 规格
    '劈开面/自然面',     # 肌理
    '外墙|旧墙改造|商业内墙',  # 应用
    '轻薄可弯曲，模仿石材/木纹/陶土质感，铺贴后与建筑浑然一体',  # 视觉
    '专用粘结剂+勾缝；阴角可弯曲包覆，无需切角',  # 构造
    '高',           # 耐久
    '≥30 年',       # 使用寿命
    '一般免维护',     # 维护
    '环球软瓷|莫干山',   # 供应商
    'flex_001_main.jpg|flex_001_detail.jpg',  # 图片
    '柔性饰面砖的防火等级、吸水率、与基层的粘结强度',  # 考试要点
    '万科城市花园外改造',   # 考试案例
    0.1,            # 考试权重
    '软瓷材料样本册 2024 版',  # 来源
    '适合曲面/异形建筑，铺贴后接缝少',  # 备注
]

for col_idx, val in enumerate(EXAMPLE, 1):
    cell = ws.cell(row=2, column=col_idx, value=val)
    cell.fill = EXMAPLE_FILL
    cell.alignment = WRAP
    cell.border = BORDER

# 列宽
widths = {
    '编码': 12, '中文名': 14, '英文名': 22, '分类代码': 14, '子分类': 16,
    '单价': 10, '单位': 8, '单价(立方)': 12, '施工费': 10, '损耗系数': 10,
    '防火等级': 10, '防火说明': 16, '造价档': 8, '环保等级': 14, '规范标准': 16,
    '环保认证': 14, '密度': 18, '强度': 20, '导热系数': 16, '吸水率': 10,
    '质感': 20, '色系': 16, '常见规格': 22, '肌理': 18, '应用场景': 24,
    '视觉效果': 36, '构造节点': 30, '耐久性': 10, '使用寿命': 12, '维护周期': 18,
    '供应商': 20, '图片文件名': 28, '考试要点': 28, '考试案例': 22, '考试权重': 10,
    '来源': 22, '备注': 28,
}
for col_idx, (name, _, _, _, _) in enumerate(COLUMNS, 1):
    w = widths.get(name, 14)
    ws.column_dimensions[get_column_letter(col_idx)].width = w
ws.row_dimensions[1].height = 28
ws.row_dimensions[2].height = 120
ws.freeze_panes = 'D2'  # 冻结前 3 列 + 表头

# ============================
# Sheet 2: 字典
# ============================
ws2 = wb.create_sheet('字典')

# 分类
ws2.cell(row=1, column=1, value='分类代码').fill = HEADER_FILL
ws2.cell(row=1, column=1).font = HEADER_FONT
ws2.cell(row=1, column=2, value='分类名称').fill = HEADER_FILL
ws2.cell(row=1, column=2).font = HEADER_FONT
ws2.cell(row=1, column=3, value='父级').fill = HEADER_FILL
ws2.cell(row=1, column=3).font = HEADER_FONT
CATS = [
    ('metal',          '金属材料',     ''),
    ('metal.steel',    '建筑钢材',     'metal'),
    ('metal.aluminum', '铝合金',       'metal'),
    ('metal.copper',   '铜/钛/锌',     'metal'),
    ('metal.stainless','不锈钢',       'metal'),
    ('concrete',       '混凝土',       ''),
    ('concrete.normal','普通混凝土',   'concrete'),
    ('concrete.hpc',   '高性能混凝土', 'concrete'),
    ('concrete.scc',   '自密实混凝土', 'concrete'),
    ('concrete.fairface','清水混凝土', 'concrete'),
    ('masonry',        '砌体材料',     ''),
    ('wood',           '木材',         ''),
    ('glass',          '玻璃',         ''),
    ('glass.float',    '浮法玻璃',     'glass'),
    ('glass.tempered', '钢化玻璃',     'glass'),
    ('glass.laminated','夹层玻璃',     'glass'),
    ('glass.low_e',    'Low-E中空',    'glass'),
    ('glass.vacuum',   '真空玻璃',     'glass'),
    ('stone',          '石材',         ''),
    ('stone.granite',  '花岗岩',       'stone'),
    ('stone.marble',   '大理石',       'stone'),
    ('stone.sandstone','砂岩',         'stone'),
    ('stone.artificial','人造石材',    'stone'),
    ('membrane',       '膜材',         ''),
    ('insulation',     '保温/防水',    ''),
    ('finishing',      '装饰材料',     ''),
    ('finishing.paint','涂料',         'finishing'),
    ('finishing.tile', '瓷砖',         'finishing'),
    ('finishing.metal','金属板',       'finishing'),
    ('finishing.wood', '木饰面',       'finishing'),
    ('finishing.glass','玻璃',         'finishing'),
    ('composite',      '复合材料',     ''),
]
for i, (code, name, parent) in enumerate(CATS, 2):
    ws2.cell(row=i, column=1, value=code).border = BORDER
    ws2.cell(row=i, column=2, value=name).border = BORDER
    ws2.cell(row=i, column=3, value=parent).border = BORDER

# 防火等级
row = len(CATS) + 4
ws2.cell(row=row, column=1, value='防火等级').fill = HEADER_FILL
ws2.cell(row=row, column=1).font = HEADER_FONT
ws2.cell(row=row, column=2, value='含义').fill = HEADER_FILL
ws2.cell(row=row, column=2).font = HEADER_FONT
for i, (k, v) in enumerate([('A1', '不燃材料'), ('B1', '难燃材料'), ('B2', '可燃材料')], 1):
    ws2.cell(row=row+i, column=1, value=k).border = BORDER
    ws2.cell(row=row+i, column=2, value=v).border = BORDER

# 造价档
row += 5
ws2.cell(row=row, column=1, value='造价档').fill = HEADER_FILL
ws2.cell(row=row, column=1).font = HEADER_FONT
ws2.cell(row=row, column=2, value='范围').fill = HEADER_FILL
ws2.cell(row=row, column=2).font = HEADER_FONT
for i, (k, v) in enumerate([('低', '<200 元/m²'), ('中', '200~500 元/m²'), ('高', '>500 元/m²')], 1):
    ws2.cell(row=row+i, column=1, value=k).border = BORDER
    ws2.cell(row=row+i, column=2, value=v).border = BORDER

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 14

# ============================
# Sheet 3: 说明
# ============================
ws3 = wb.create_sheet('说明')
ws3.cell(row=1, column=1, value='材料表录入说明').font = Font(bold=True, size=14, color='FF9E4A')
INSTRUCTIONS = [
    '',
    '【必填项】只有 3 个：编码、中文名、分类代码。其他都可以留空。',
    '',
    '【分类代码】必须从"字典"sheet 里选（如 masonry / metal / stone.granite）。',
    '   不在表里的会自动跳过。',
    '',
    '【多值字段】用 | 分隔，例如：',
    '   质感: 哑光|粗糙|自然纹理',
    '   应用场景: 外墙|幕墙|室内',
    '   图片文件名: flex_001_main.jpg|flex_001_detail.jpg',
    '',
    '【图片】放文件名到 D:\\Mac\\Mac\\workteam\\05_space\\03_architect\\_ArchitectLib\\MaterialDb\\media\\images\\',
    '   填到"图片文件名"列，多个用 | 分隔。前端会自动显示在卡片背景。',
    '',
    '【编码】要唯一。如果导入时编码已存在：',
    '   - 默认：跳过该行（提示）',
    '   - 加 --overwrite 参数：覆盖更新',
    '   - 加 --strict 参数：报错退出',
    '',
    '【使用流程】',
    '   1. 复制本模板另存一份（不要改原模板）',
    '   2. 在"材料"sheet 第 3 行起开始填新材质（不要改第 2 行示例）',
    '   3. 保存文件',
    '   4. 运行: python import_materials.py 你的文件名.xlsx',
    '   5. 刷新 MaterialWeb 页面看新材质',
    '',
    '【干跑模式】只检查不写库:',
    '   python import_materials.py 你的文件名.xlsx --dry-run',
    '',
    '【其他】',
    '   - 单元格有红框表示必填留空了',
    '   - 列名鼠标悬停会显示说明 comment',
    '   - 任何列名拼写要和模板完全一致',
]
for i, line in enumerate(INSTRUCTIONS, 2):
    ws3.cell(row=i, column=1, value=line).alignment = WRAP

ws3.column_dimensions['A'].width = 90

# 保存
wb.save(OUT)
print(f'✓ 已生成: {OUT}')
print(f'  路径: {os.path.abspath(OUT)}')
print(f'  示例行: 软瓷片 (FLEX_001, masonry 类)')
