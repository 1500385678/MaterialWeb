# -*- coding: utf-8 -*-
"""_frame_to_plan.py · Frame 待开发功能1 → Plan 下一天
规划 Agent 在 22:00 cron 调用
"""
import io
import sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import argparse
import os
import shutil
from datetime import date, timedelta
from openpyxl import load_workbook

# 自动定位:脚本在 canvasweb/.AutoEvolution/ 下时,ROOT 是 canvasweb/
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FRAME = os.path.join(ROOT, '.AutoEvolution', '.CanvasWebFrame.xlsx')
PLAN = os.path.join(ROOT, '.AutoEvolution', '.CanvasWebPlan-8月.xlsx')

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side





# Frame 工作集
SHEET_BACKLOG = '待开发功能1'
SHEET_DONE = '核心(天天用)'

# 编码规则
PREFIX = 'D'

# Plan sheet 命名规则
WEEKDAY_CN = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']


def ensure_frame_columns(ws):
    """确保 G(状态) H(进 Plan 日) 列存在"""
    if ws.cell(1, 7).value is None or '状态' not in str(ws.cell(2, 7).value or ''):
        ws.cell(1, 7, '元数据')
        ws.cell(2, 7, '状态')
        ws.cell(2, 8, '进 Plan 日')
    # 给所有行默认 G="未开发"
    for row in range(3, ws.max_row + 1):
        if ws.cell(row, 1).value or ws.cell(row, 2).value:
            if not ws.cell(row, 7).value:
                ws.cell(row, 7, '未开发')


def get_next_d_number(ws, used=None):
    """返回下一个可用 D-XXX 编号(支持外部传入 used 避免重复)"""
    if used is None:
        used = set()
        for row in range(3, ws.max_row + 1):
            v = str(ws.cell(row, 1).value or '')
            if v.startswith(f'[{PREFIX}-'):
                try:
                    n = int(v[3:].rstrip(']'))
                    used.add(n)
                except ValueError:
                    pass
    n = 1
    while n in used:
        n += 1
    used.add(n)
    return f'{PREFIX}-{n:03d}', used


def pick_today_plan_sheet(wb, target_date):
    """找 Plan 里对应 target_date 的 Day sheet 名"""
    weekday = WEEKDAY_CN[target_date.weekday()]
    for sn in wb.sheetnames:
        if weekday in sn and sn.startswith('Day'):
            return sn
    # 兜底:第一个 Day sheet
    for sn in wb.sheetnames:
        if sn.startswith('Day'):
            return sn
    return wb.sheetnames[0] if wb.sheetnames else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--max', type=int, default=3, help='每天最多挑几个任务')
    ap.add_argument('--dry-run', action='store_true', help='只打印不写')
    ap.add_argument('--date', help='写进 Plan 哪个日期(默认明天)')
    args = ap.parse_args()

    target_date = date.fromisoformat(args.date) if args.date else date.today() + timedelta(days=1)
    today_iso = date.today().isoformat()

    print(f'目标日期: {target_date} ({WEEKDAY_CN[target_date.weekday()]})')
    print(f'挑任务上限: {args.max}')

    # 1. 读 Frame
    wb_f = load_workbook(FRAME)
    if SHEET_BACKLOG not in wb_f.sheetnames:
        print(f'❌ Frame 缺 {SHEET_BACKLOG} sheet')
        return
    ws_back = wb_f[SHEET_BACKLOG]
    ensure_frame_columns(ws_back)

    # 2. 收集"未开发"且无进 Plan 日的任务
    todo = []
    for row in range(3, ws_back.max_row + 1):
        name = ws_back.cell(row, 2).value
        desc = ws_back.cell(row, 3).value
        status = ws_back.cell(row, 7).value
        plan_date = ws_back.cell(row, 8).value
        if name and (status == '未开发' or not status) and not plan_date:
            todo.append((row, str(name).strip(), str(desc or '').strip()))

    print(f'候选: {len(todo)} 项')

    if not todo:
        print('✅ 今日无新任务可规划')
        return

    # 3. 挑 N 个 + 分配 D-XXX 编码
    picks = todo[:args.max]
    new_rows = []
    used = None
    for row, name, desc in picks:
        code, used = get_next_d_number(ws_back, used)
        new_rows.append((row, code, name, desc))
        print(f'  {code} · {name}')

    if args.dry_run:
        print('\n[DRY-RUN] 未写盘')
        return

    # 4. 写 Plan
    wb_p = load_workbook(PLAN)
    day_sn = pick_today_plan_sheet(wb_p, target_date)
    if not day_sn:
        print(f'❌ Plan 无 sheet')
        return
    ws_day = wb_p[day_sn]

    # 找下一个空行
    start_row = 4
    while ws_day.cell(start_row, 1).value:
        start_row += 1

    for i, (row, code, name, desc) in enumerate(new_rows):
        r = start_row + i
        ws_day.cell(r, 1, name)
        ws_day.cell(r, 2, f'[{code}]')
        ws_day.cell(r, 3, '8h')  # 默认估算
        ws_day.cell(r, 4, 'medium')  # 默认风险
        ws_day.cell(r, 5, 'TBD')  # 涉及文件
        ws_day.cell(r, 6, desc or '见 Frame 描述')  # 验收标准
        ws_day.cell(r, 7, 'todo')  # 状态(新加列)
        print(f'  Plan {day_sn} row {r}: [{code}] {name}')

    wb_p.save(PLAN)

    # 5. 更新 Frame G/H
    for row, code, name, desc in new_rows:
        ws_back.cell(row, 1, f'[{code}]')
        ws_back.cell(row, 7, '计划中')
        ws_back.cell(row, 8, today_iso)

    wb_f.save(FRAME)
    print(f'\n✅ Frame+Plan 已更新: {len(new_rows)} 项进 Plan')


if __name__ == '__main__':
    main()
