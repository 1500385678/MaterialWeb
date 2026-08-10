# -*- coding: utf-8 -*-
"""_frame_renumber.py · 重新编号,消除漂移
任何 sheet 增删行后调用
"""
import io
import sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import argparse
import os
import re
from openpyxl import load_workbook

# 自动定位:脚本在 canvasweb/.AutoEvolution/ 下时,ROOT 是 canvasweb/
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FRAME = os.path.join(ROOT, '.AutoEvolution', '.CanvasWebFrame.xlsx')
PLAN = os.path.join(ROOT, '.AutoEvolution', '.CanvasWebPlan-8月.xlsx')





# 编码前缀 → sheet 名
PREFIX_TO_SHEET = {
    'A': '核心(天天用)',
    'B': '辅助(经常用)',
    'C': '实验(新功能)',
    'D': '待开发功能1',
    'E': '待开发功能2',
    'F': '待开发功能3',
}

CODE_PATTERN = re.compile(r'^\[([A-F])-(\d{3})\]')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    wb = load_workbook(FRAME)
    changes = []

    for prefix, sn in PREFIX_TO_SHEET.items():
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        n = 1
        for row in range(3, ws.max_row + 1):
            v = ws.cell(row, 1).value
            if not v:
                continue
            m = CODE_PATTERN.match(str(v).strip())
            if not m:
                continue
            old = str(v).strip()
            new = f'[{prefix}-{n:03d}]'
            if old != new:
                ws.cell(row, 1).value = new
                changes.append(f'  {sn} row {row}: {old} → {new}')
            n += 1

    if not changes:
        print('✅ 无需重编号')
        return

    print(f'需重编号: {len(changes)} 处')
    for c in changes[:20]:
        print(c)
    if len(changes) > 20:
        print(f'  ...还有 {len(changes) - 20} 处')

    if args.dry_run:
        print('\n[DRY-RUN] 未写盘')
        return

    wb.save(FRAME)
    print(f'\n✅ 已重编号: {len(changes)} 处')


if __name__ == '__main__':
    main()
