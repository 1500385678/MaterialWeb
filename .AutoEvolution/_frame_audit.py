# -*- coding: utf-8 -*-
"""_frame_audit.py · 漂移巡检:Frame vs Plan vs GitHub Issues
周日晚 cron + 手动调用
"""
import io
import sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import argparse
import urllib.request
import urllib.error
import json
import os
from openpyxl import load_workbook

# 自动定位:脚本在 canvasweb/.AutoEvolution/ 下时,ROOT 是 canvasweb/
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FRAME = os.path.join(ROOT, '.AutoEvolution', '.CanvasWebFrame.xlsx')
PLAN = os.path.join(ROOT, '.AutoEvolution', '.CanvasWebPlan-8月.xlsx')






GH_API = 'https://api.github.com'
REPO = '1500385678/canvasweb'


def get_gh_issues():
    """拉所有 open + closed issues"""
    token = os.environ.get('GH_TOKEN', '')
    headers = {'Accept': 'application/vnd.github+json', 'User-Agent': 'canvasweb-audit'}
    if token:
        headers['Authorization'] = f'token {token}'

    issues = []
    page = 1
    while True:
        url = f'{GH_API}/repos/{REPO}/issues?state=all&per_page=100&page={page}'
        try:
            req = urllib.request.Request(url, headers=headers)
            # 绕过系统代理
            opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
            resp = opener.open(req, timeout=15)
            data = json.loads(resp.read())
        except Exception as e:
            print(f'⚠️  GitHub 拉取失败: {e}')
            return []
        if not data:
            break
        for item in data:
            if 'pull_request' in item:
                continue
            issues.append({
                'number': item['number'],
                'title': item['title'],
                'state': item['state'],
            })
        if len(data) < 100:
            break
        page += 1
    return issues


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--no-github', action='store_true', help='跳过 GitHub 检查')
    ap.add_argument('--grace-days', type=int, default=2, help='宽限期(未进 Plan 几天内不算漂移,默认 2)')
    args = ap.parse_args()

    from datetime import date, timedelta
    today = date.today()
    grace_start = today - timedelta(days=args.grace_days)

    issues_found = []
    stats = {
        'frame_total': 0,
        'frame_by_status': {},
        'frame_未进_plan_宽限内': 0,
        'frame_未进_plan_超期': 0,
    }

    # 1. Frame 检查
    wb_f = load_workbook(FRAME)
    todo_in_frame = []
    for sn in ['待开发功能1', '待开发功能2', '待开发功能3', '核心(天天用)', '辅助(经常用)', '实验(新功能)']:
        if sn not in wb_f.sheetnames:
            continue
        ws = wb_f[sn]
        for row in range(3, ws.max_row + 1):
            code = str(ws.cell(row, 1).value or '').strip('[]')
            name = ws.cell(row, 2).value
            if not code or not name:
                continue
            status = ws.cell(row, 7).value or '未开发'
            plan_date = str(ws.cell(row, 8).value or '').strip()
            todo_in_frame.append((sn, code, str(name).strip(), status, plan_date))
            stats['frame_total'] += 1
            stats['frame_by_status'][status] = stats['frame_by_status'].get(status, 0) + 1

    # 找漂移:未开发 + 无 plan_date + 超过宽限期
    for sn, code, name, status, plan_date in todo_in_frame:
        if sn == '待开发功能1' and status == '未开发' and not plan_date:
            # 还没进 Plan,判断是否超期
            # 没日期信息?默认当成宽限内(新加的)
            stats['frame_未进_plan_宽限内'] += 1

    # 找漂移:Frame 标"计划中"但 Plan 无(进了 Plan 但 Plan 找不到)
    # (后面会用到 plan_codes)

    # 2. Plan 检查
    wb_p = load_workbook(PLAN)
    plan_codes = set()
    plan_done_codes = set()
    for sn in wb_p.sheetnames:
        if not sn.startswith('Day'):
            continue
        ws = wb_p[sn]
        for row in range(4, ws.max_row + 1):
            code = str(ws.cell(row, 2).value or '').strip('[]')
            status = ws.cell(row, 7).value
            if not code:
                continue
            plan_codes.add(code)
            if status == 'done':
                plan_done_codes.add(code)

    # 找漂移:Frame 中"计划中" + plan_date 有值,但 Plan 找不到 → 进了 Plan 但 Plan 找不到
    for sn, code, name, status, plan_date in todo_in_frame:
        if status == '计划中' and plan_date and code not in plan_codes:
            issues_found.append(f'  📤 {sn} {code} · {name} · Frame 标"计划中"但 Plan 无')

    # 找漂移:Plan 标 done 但 Frame 还在"待开发1"
    frame_codes = {c for _, c, _, _, _ in todo_in_frame}
    for code in plan_done_codes:
        # 找原属 sheet
        for sn, c, name, status, plan_date in todo_in_frame:
            if c == code and sn == '待开发功能1':
                issues_found.append(f'  🔁 {code} · {name} · Plan 标 done 但 Frame 还在待开发1 · 该调 _plan_to_frame.py')

    # 3. GitHub 检查
    if not args.no_github:
        print('拉 GitHub issues...')
        issues = get_gh_issues()
        print(f'  GitHub issues: {len(issues)}')
        for iss in issues:
            t = iss['title']
            # 提取 [D-001] 类编码
            if t.startswith('[') and ']' in t:
                code = t[1:t.index(']')].strip()
                # Frame 标已完成但 issue 还 open
                for sn, c, name, status, plan_date in todo_in_frame:
                    if c == code and status == '已完成' and iss['state'] == 'open':
                        issues_found.append(f'  🐙 GitHub #{iss["number"]} {code} · 标题={t} · Frame 已完成但 issue 还 open')
                # issue 关闭了但 Frame 还"未开发"或"计划中"
                for sn, c, name, status, plan_date in todo_in_frame:
                    if c == code and iss['state'] == 'closed' and status in ('未开发', '计划中'):
                        issues_found.append(f'  🐙 GitHub #{iss["number"]} {code} · 标题={t} · Issue 已关但 Frame 还 {status}')

    # 4. 输出
    print(f'\n=== 统计 ===')
    print(f'Frame 总数: {stats["frame_total"]}')
    for k, v in stats['frame_by_status'].items():
        print(f'  {k}: {v}')
    print(f'  未进 Plan(宽限内): {stats["frame_未进_plan_宽限内"]}')
    print(f'  未进 Plan(超期): {stats["frame_未进_plan_超期"]}')

    if not issues_found:
        print(f'\n✅ 一切正常,无漂移')
        return

    print(f'\n⚠️  发现 {len(issues_found)} 处漂移:')
    for iss in issues_found:
        print(iss)
    sys.exit(1)


if __name__ == '__main__':
    main()
