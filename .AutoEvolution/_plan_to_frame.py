# -*- coding: utf-8 -*-
"""_plan_to_frame.py · Plan 完成 → Frame 状态 + 工作集迁移
执行 Agent 在任务完成后调用
"""
import io
import sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import argparse
import os
from datetime import date
from openpyxl import load_workbook

# 自动定位:脚本在 canvasweb/.AutoEvolution/ 下时,ROOT 是 canvasweb/
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FRAME = os.path.join(ROOT, '.AutoEvolution', '.CanvasWebFrame.xlsx')
PLAN = os.path.join(ROOT, '.AutoEvolution', '.CanvasWebPlan-8月.xlsx')






SHEET_BACKLOG = '待开发功能1'
SHEET_DONE = '核心(天天用)'
DONE_PREFIX = 'A'  # 核心用 A-XXX 编码


def get_next_a_number(ws, used=None):
    """返回"核心"sheet 下一个可用 A-XXX"""
    if used is None:
        used = set()
        for row in range(3, ws.max_row + 1):
            v = str(ws.cell(row, 1).value or '').strip('[]')
            if v.startswith(f'{DONE_PREFIX}-'):
                try:
                    n = int(v[3:])
                    used.add(n)
                except ValueError:
                    pass
    n = 1
    while n in used:
        n += 1
    used.add(n)
    return f'[{DONE_PREFIX}-{n:03d}]', used


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    # 1. 读 Plan, 找 done
    wb_p = load_workbook(PLAN)
    done_tasks = []  # (day_sheet, row, code, name)
    for sn in wb_p.sheetnames:
        if not sn.startswith('Day'):
            continue
        ws = wb_p[sn]
        for row in range(4, ws.max_row + 1):
            status = ws.cell(row, 7).value
            if status == 'done':
                name = ws.cell(row, 1).value
                code = str(ws.cell(row, 2).value or '').strip('[]')
                if code.startswith('D-') and name:
                    done_tasks.append((sn, row, code, str(name).strip()))

    print(f'Plan 中已完成任务: {len(done_tasks)}')

    if not done_tasks:
        print('✅ 无新完成')
        return

    # 2. 读 Frame
    wb_f = load_workbook(FRAME)
    ws_back = wb_f[SHEET_BACKLOG]
    ws_done = wb_f[SHEET_DONE]

    # 3. 对每个 done:从待开发 1 删,在核心加
    used_a = None
    for day_sn, plan_row, code, name in done_tasks:
        # 找 Frame 中对应行
        target_row = None
        for r in range(3, ws_back.max_row + 1):
            v = str(ws_back.cell(r, 1).value or '').strip('[]')
            if v == code:
                target_row = r
                break
        if not target_row:
            print(f'  ⚠️  {code} 在 Frame {SHEET_BACKLOG} 找不到,跳过')
            continue

        desc = ws_back.cell(target_row, 3).value
        # 移走
        ws_back.delete_rows(target_row, 1)
        # 加到"核心",分配 A-XXX
        new_row = 3
        while ws_done.cell(new_row, 1).value:
            new_row += 1
        new_code, used_a = get_next_a_number(ws_done, used_a)
        ws_done.cell(new_row, 1, new_code)
        ws_done.cell(new_row, 2, name)
        ws_done.cell(new_row, 3, desc or '')
        ws_done.cell(new_row, 7, '已完成')
        ws_done.cell(new_row, 8, date.today().isoformat())
        print(f'  ✅ {code} · {name} → {SHEET_DONE} {new_code}')

    if args.dry_run:
        print('\n[DRY-RUN] 未写盘')
        return

    wb_f.save(FRAME)
    # Plan 状态保留(作为历史),不改
    print(f'\n✅ Frame 已更新: {len(done_tasks)} 项归档到核心')


if __name__ == '__main__':
    main()
