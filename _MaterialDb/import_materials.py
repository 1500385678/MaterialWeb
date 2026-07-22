"""
从 Excel 导入材料到 MaterialDB
用法:
  python import_materials.py 材料表.xlsx
  python import_materials.py 材料表.xlsx --dry-run      # 只校验不写库
  python import_materials.py 材料表.xlsx --overwrite    # 编码冲突时覆盖
  python import_materials.py 材料表.xlsx --strict        # 冲突时直接报错
"""
import sys
import os
import json
import sqlite3
import argparse
from datetime import datetime
from openpyxl import load_workbook

DB = r'D:\Mac\Mac\workteam\05_space\03_architect\_ArchitectLib\MaterialDb\materials.db'

# 字段映射：Excel 列名 → (DB 字段, 类型)
# 类型: 'str'=TEXT, 'num'=REAL, 'list'=JSON array, 'json'=JSON
COLUMN_MAP = {
    '编码':         ('code',             'str'),
    '中文名':       ('name_cn',          'str'),
    '英文名':       ('name_en',          'str'),
    '分类代码':     ('category_code',    'str'),
    '子分类':       ('sub_category',     'str'),
    '单价':         ('unit_price',       'num'),
    '单位':         ('unit',             'str'),
    '单价(立方)':   ('unit_price_m3',    'num'),
    '施工费':       ('labor_cost',       'num'),
    '损耗系数':     ('loss_factor',      'num'),
    '防火等级':     ('fire_rating',      'str'),
    '防火说明':     ('fire_note',        'str'),
    '造价档':       ('cost_tier',        'str'),
    '环保等级':     ('env_grade',        'str'),
    '规范标准':     ('std_code',         'str'),
    '环保认证':     ('eco_cert',         'str'),
    '密度':         ('density',          'str'),
    '强度':         ('strength',         'str'),
    '导热系数':     ('thermal_cond',     'str'),
    '吸水率':       ('water_absorp',     'str'),
    '质感':         ('texture',          'list'),
    '色系':         ('color_series',     'list'),
    '常见规格':     ('specs',            'str'),
    '肌理':         ('patterns',         'str'),
    '应用场景':     ('applications_json','list'),
    '视觉效果':     ('visual_desc',      'str'),
    '构造节点':     ('structure_notes',  'str'),
    '耐久性':       ('durability',       'str'),
    '使用寿命':     ('lifespan_years',   'str'),
    '维护周期':     ('maintenance',      'str'),
    '供应商':       ('suppliers_json',   'str'),
    '图片文件名':   ('image_urls',       'list'),
    '考试要点':     ('exam_points',      'list'),
    '考试案例':     ('exam_cases',       'str'),
    '考试权重':     ('exam_weight',      'num'),
    '来源':         ('source_doc',       'str'),
    '备注':         ('remark',           'str'),
}

REQUIRED = {'编码', '中文名', '分类代码'}

# JSON 字段（list 类型里哪些进 JSON 数组）
JSON_FIELDS = {
    'texture', 'color_series', 'applications', 'image_urls', 'exam_points'
}


def to_str(v):
    if v is None: return ''
    return str(v).strip()


def to_num(v, default=None):
    if v is None or v == '': return default
    try: return float(v)
    except: return default


def to_list(v):
    """把 'a|b|c' 拆成 ['a','b','c']，去空"""
    if v is None: return []
    s = str(v).strip()
    if not s: return []
    return [x.strip() for x in s.replace('，', '|').split('|') if x.strip()]


def to_json(v):
    """把 list 转成 JSON 字符串存库"""
    arr = to_list(v)
    return json.dumps(arr, ensure_ascii=False)


def parse_row(row, headers):
    """把一行 Excel 数据解析成 {db_field: value}"""
    out = {}
    for col_idx, header in enumerate(headers):
        if header not in COLUMN_MAP: continue
        db_field, ftype = COLUMN_MAP[header]
        raw = row[col_idx] if col_idx < len(row) else None
        if ftype == 'str':
            out[db_field] = to_str(raw)
        elif ftype == 'num':
            out[db_field] = to_num(raw)
        elif ftype == 'list':
            if db_field in JSON_FIELDS:
                out[db_field] = to_json(raw) if to_str(raw) else '[]'
            else:
                out[db_field] = ', '.join(to_list(raw))
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx', help='Excel 文件路径')
    ap.add_argument('--dry-run', action='store_true', help='只校验不写库')
    ap.add_argument('--overwrite', action='store_true', help='编码冲突时覆盖')
    ap.add_argument('--strict', action='store_true', help='冲突时报错退出')
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        print(f'✗ 文件不存在: {args.xlsx}')
        sys.exit(1)
    if not os.path.exists(DB):
        print(f'✗ 数据库不存在: {DB}')
        sys.exit(1)

    print(f'>>> 读取: {args.xlsx}')
    wb = load_workbook(args.xlsx, data_only=True)
    if '材料' not in wb.sheetnames:
        print('✗ Excel 里找不到"材料"sheet')
        sys.exit(1)
    ws = wb['材料']
    headers = [c.value for c in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # 过滤：跳过空行
    rows = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
    print(f'>>> 数据行数: {len(rows)}')

    # 校验
    errors = []
    parsed = []
    for i, row in enumerate(rows, start=2):  # row 1 is header
        rec = {'_row': i, '_raw': row}
        # 必填校验
        for req in REQUIRED:
            col_idx = headers.index(req) if req in headers else -1
            if col_idx < 0 or col_idx >= len(row) or to_str(row[col_idx]) == '':
                errors.append(f'第 {i} 行: 缺少必填字段 [{req}]')
        rec.update(parse_row(row, headers))
        parsed.append(rec)

    # 查分类
    conn = sqlite3.connect(DB)
    cat_map = {}
    for code, name in conn.execute('SELECT code, name FROM categories').fetchall():
        cat_map[code] = name
    for r in parsed:
        cat_code = r.get('category_code', '')
        if cat_code and cat_code not in cat_map:
            errors.append(f"第 {r['_row']} 行: 分类代码 '{cat_code}' 不存在（看'字典'sheet）")
        r['_category_name'] = cat_map.get(cat_code, '')

    if errors:
        print('\n=== ✗ 校验失败 ===')
        for e in errors: print('  ' + e)
        conn.close()
        sys.exit(1)

    print('>>> 校验通过')

    if args.dry_run:
        print('\n=== 干跑模式 — 不会写库 ===')
        for r in parsed:
            print(f'  [预览] 第 {r["_row"]} 行: {r.get("code")} | {r.get("name_cn")} | {r.get("_category_name")}')
        conn.close()
        return

    # 写入
    inserted, updated, skipped = 0, 0, 0
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for r in parsed:
        # 找 category_id
        cat_id = conn.execute(
            'SELECT id FROM categories WHERE code = ?', (r['category_code'],)
        ).fetchone()
        cat_id = cat_id[0] if cat_id else None

        # 检查重复
        existing = conn.execute(
            'SELECT id FROM materials WHERE code = ?', (r['code'],)
        ).fetchone()

        # 准备数据
        data = dict(r)
        data.pop('_row', None); data.pop('_raw', None); data.pop('_category_name', None)
        data.pop('category_code', None)  # 临时字段，不入表
        data['category_id'] = cat_id
        data['status'] = 'active'
        if not data.get('created_at'):
            data['created_at'] = now
        data['updated_at'] = now

        if existing:
            if args.strict:
                print(f'✗ 编码 {r["code"]} 已存在 (--strict 模式)')
                conn.close()
                sys.exit(1)
            elif args.overwrite:
                # 更新
                cols = [k for k in data.keys() if k != 'created_at']
                sets = ', '.join(f'{k} = ?' for k in cols)
                vals = [data[k] for k in cols] + [existing[0]]
                conn.execute(f'UPDATE materials SET {sets} WHERE id = ?', vals)
                updated += 1
                print(f'  ↻ 更新: {r["code"]} ({r["name_cn"]})')
            else:
                skipped += 1
                print(f'  ⊘ 跳过: {r["code"]} ({r["name_cn"]}) — 已存在（用 --overwrite 覆盖）')
        else:
            cols = list(data.keys())
            placeholders = ', '.join('?' for _ in cols)
            vals = [data[k] for k in cols]
            conn.execute(f'INSERT INTO materials ({", ".join(cols)}) VALUES ({placeholders})', vals)
            inserted += 1
            print(f'  ✓ 新增: {r["code"]} ({r["name_cn"]}) → {r["_category_name"]}')

    conn.commit()
    conn.close()

    print(f'\n=== 完成 ===')
    print(f'  ✓ 新增: {inserted}')
    print(f'  ↻ 更新: {updated}')
    print(f'  ⊘ 跳过: {skipped}')
    print(f'  刷新 http://127.0.0.1:5188/ 查看新材质')


if __name__ == '__main__':
    main()
